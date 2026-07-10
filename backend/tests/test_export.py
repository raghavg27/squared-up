"""Group spreadsheet export (`GET /groups/<id>/export`).

Verifies the endpoint streams a real .xlsx and that its per-member columns
reconcile to the live group balances — the bold ``Total balance`` footer — so
the ledger is internally consistent (Spec §6 / I3). Outsiders get 404.
"""

import io
import uuid

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from core.auth import issue_tokens
from core.models import Settlement, User

pytestmark = pytest.mark.django_db


@pytest.fixture
def client():
    return APIClient()


def _mk_user(name):
    return User.objects.create(name=name, locale="en").id


def _as(client, user_id):
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {issue_tokens(user_id)['access']}")
    return client


def _make_group_with_history(client, a, b, c):
    from django.utils import timezone

    gid = client.post(
        "/api/v1/groups",
        {"name": "Goa [2024]", "type": "trip", "member_ids": [b, c]},
        format="json",
    ).data["id"]
    client.post(
        "/api/v1/expenses",
        {
            "group_id": gid,
            "description": "Dinner",
            "amount_paise": 30000,
            "expense_date": "2024-01-10",
            "payers": [{"user_id": a, "paid_paise": 30000}],
            "split": {"type": "equal", "participants": [a, b, c]},
        },
        format="json",
        HTTP_IDEMPOTENCY_KEY=str(uuid.uuid4()),
    )
    # Confirmed settlement Bhavna → Aarav ₹100 (moves both toward zero).
    Settlement.objects.create(
        group_id=gid, from_user_id=b, to_user_id=a, amount_paise=10000,
        method="upi", status="confirmed", confirmed_at=timezone.now(),
    )
    return gid


def test_group_export_xlsx_reconciles(client):
    a, b, c = _mk_user("Aarav"), _mk_user("Bhavna"), _mk_user("Chetan")
    _as(client, a)
    gid = _make_group_with_history(client, a, b, c)

    r = client.get(f"/api/v1/groups/{gid}/export")
    assert r.status_code == 200
    assert r["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in r["Content-Disposition"]

    rows = list(load_workbook(io.BytesIO(r.content)).active.iter_rows(values_only=True))
    header, total = rows[0], rows[-1]
    assert header[:5] == ("Date", "Description", "Category", "Cost", "Currency")
    names = header[5:]
    assert set(names) == {"Aarav", "Bhavna", "Chetan"}
    assert total[1] == "Total balance"

    # Each member column sums (over data rows) to its Total balance cell, and
    # the balances net to zero across the group (I3).
    for ci in range(5, len(header)):
        col = sum(rows[k][ci] or 0 for k in range(1, len(rows) - 1))
        assert col == total[ci]
    assert sum(v or 0 for v in total[5:]) == 0
    by_name = {names[i]: total[5 + i] for i in range(len(names))}
    assert by_name == {"Aarav": 100, "Bhavna": 0, "Chetan": -100}


def test_group_export_outsider_404(client):
    a, d = _mk_user("Owner"), _mk_user("Outsider")
    _as(client, a)
    gid = client.post(
        "/api/v1/groups", {"name": "Trip", "type": "trip", "member_ids": []}, format="json"
    ).data["id"]

    _as(client, d)
    r = client.get(f"/api/v1/groups/{gid}/export")
    assert r.status_code == 404
    assert r.data["error"]["code"] == "NOT_FOUND"
