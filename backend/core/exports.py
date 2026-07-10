"""Group expense export → a real ``.xlsx`` spreadsheet (Splitwise-style ledger).

One row per expense and per confirmed settlement, in chronological order. There
is a column per member holding their signed net for that line
(``paid − owed``: +ve = they lent, −ve = they borrowed), so a bold
``Total balance`` footer whose column sums equal the group's live balances
(Spec §6). Paise → rupees conversion happens **only here**, at the presentation
boundary — the same rule the frontend's ``format.ts`` follows. Values are
written as real numbers with a money format so recipients can sum/pivot them.
"""

from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import services
from .models import Expense, Settlement, User

# Excel numeric format — keeps money as a number (not text) so it stays sum-able.
_MONEY_FMT = "#,##0.00"
_DATE_FMT = "yyyy-mm-dd"
_FIXED_COLS = 5  # Date, Description, Category, Cost, Currency (member cols follow)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FONT = Font(bold=True)
_TOP_BORDER = Border(top=Side(style="thin", color="9CA3AF"))
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _rupees(paise: int) -> float:
    """Presentation-only paise→rupees. Integer paise ÷ 100 is exact to 2dp for
    every realistic amount; Excel then renders/sums it as a number."""
    return paise / 100


def _sanitize_sheet_title(name: str) -> str:
    """Sheet titles can't contain ``[]:*?/\\`` and cap at 31 chars (openpyxl)."""
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "Expenses"
    return cleaned[:31]


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]', " ", name).strip() or "group"
    cleaned = re.sub(r"\s+", " ", cleaned)
    return f"{cleaned} - Squared Up.xlsx"


def _ledger_rows(
    group_id: int, base_currency: str
) -> tuple[list[dict], set[int]]:
    """Build sorted ledger lines and the set of user ids that appear in them.

    Each line is ``{date, description, category, currency, cost_paise,
    nets: {uid: paise}}``. Expenses use per-share ``paid − owed``; a confirmed
    settlement credits the payer (+amount) and debits the receiver (−amount),
    mirroring ``domain.compute_nets`` so the footer reconciles exactly (§6)."""
    lines: list[dict] = []
    seen: set[int] = set()

    expenses = (
        Expense.objects.filter(group_id=group_id, deleted_at__isnull=True)
        .select_related("category")
        .prefetch_related("shares")
    )
    for e in expenses:
        nets: dict[int, int] = {}
        for s in e.shares.all():
            nets[s.user_id] = s.paid_paise - s.owed_paise
            seen.add(s.user_id)
        lines.append(
            {
                "sort": (e.expense_date, 0, e.id),
                "date": e.expense_date,
                "description": e.description,
                "category": e.category.name if e.category else "",
                "currency": e.currency,
                "cost_paise": e.amount_paise,
                "nets": nets,
            }
        )

    settlements = Settlement.objects.filter(
        group_id=group_id, deleted_at__isnull=True, status="confirmed"
    )
    for st in settlements:
        seen.update({st.from_user_id, st.to_user_id})
        lines.append(
            {
                "sort": (st.created_at.date(), 1, st.id),
                "date": st.created_at.date(),
                "description": "Payment",
                "category": "Payment",
                "currency": base_currency,
                "cost_paise": st.amount_paise,
                "nets": {
                    st.from_user_id: st.amount_paise,
                    st.to_user_id: -st.amount_paise,
                },
            }
        )

    lines.sort(key=lambda row: row["sort"])
    return lines, seen


def _member_order(group_id: int, seen: set[int]) -> list[int]:
    """Column order: active members first (as the group lists them), then any
    former member who still has history in the ledger. Both alphabetised by name
    so the sheet is stable and readable."""
    active = list(
        services.active_members(group_id)
        .order_by("user_id")
        .values_list("user_id", flat=True)
    )
    former = [uid for uid in seen if uid not in set(active)]
    names = _name_map(set(active) | seen)
    active.sort(key=lambda uid: names.get(uid, "").lower())
    former.sort(key=lambda uid: names.get(uid, "").lower())
    return active + former


def _name_map(ids: set[int]) -> dict[int, str]:
    return {
        u.id: u.name for u in User.objects.filter(id__in=ids).only("id", "name")
    }


def build_group_xlsx(group_id: int, actor_id: int) -> tuple[bytes, str]:
    """Return ``(xlsx_bytes, filename)`` for a group's full expense ledger.

    Raises ``DomainError('NOT_FOUND')`` for non-members (→ 404, no id probing);
    archived groups are exportable (read-only history, like Splitwise)."""
    group = services.require_group_member(group_id, actor_id, allow_archived=True)

    lines, seen = _ledger_rows(group_id, group.base_currency)
    members = _member_order(group_id, seen)
    names = _name_map(set(members))

    balances = services.group_balances(group_id)
    totals = {m["user_id"]: m["net_paise"] for m in balances["members"]}

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_title(group.name)

    header = ["Date", "Description", "Category", "Cost", "Currency"] + [
        names.get(uid, f"User {uid}") for uid in members
    ]
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center" if col >= _FIXED_COLS - 1 else "left")

    for line in lines:
        row = [
            line["date"],
            line["description"],
            line["category"],
            _rupees(line["cost_paise"]),
            line["currency"],
        ] + [_rupees(line["nets"].get(uid, 0)) for uid in members]
        ws.append(row)
        _style_row(ws, ws.max_row, members)

    # Footer: live group balances (§6). Equals each member column's sum.
    total_row = ["", "Total balance", "", "", group.base_currency] + [
        _rupees(totals.get(uid, 0)) for uid in members
    ]
    ws.append(total_row)
    r = ws.max_row
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = _TOTAL_FONT
        cell.border = _TOP_BORDER
        if col == 4 or col > _FIXED_COLS:
            cell.number_format = _MONEY_FMT

    _size_columns(ws, header)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), _sanitize_filename(group.name)


def _style_row(ws, r: int, members: list[int]) -> None:
    ws.cell(row=r, column=1).number_format = _DATE_FMT
    ws.cell(row=r, column=4).number_format = _MONEY_FMT
    for i in range(len(members)):
        ws.cell(row=r, column=_FIXED_COLS + 1 + i).number_format = _MONEY_FMT


def _size_columns(ws, header: list[str]) -> None:
    widths = [12, 30, 16, 12, 9] + [max(12, len(h) + 2) for h in header[_FIXED_COLS:]]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
